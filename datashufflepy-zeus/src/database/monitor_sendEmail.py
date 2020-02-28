# -*- coding: utf-8 -*-
import os
import sys
curPath = os.path.abspath(os.path.dirname(__file__))
sys.path.append(curPath[:-9])
import datetime
import time
from database._phoenix_hbase import *
from log.data_log import Logger
import openpyxl
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import smtplib

phoenix = PhoenixHbase('')
logger = Logger().logger


sender = 'wenzhong.ding@pactera.com'
authCode = 'gUZ4h@@6'

def sendEmailMessage(email, message,date):
    print("sendEmailMessage, message:" + message)
    logger.info("sendEmailMessage, message:" + message)
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = email
    msg['Subject'] = Header(f'数据清洗统计{date}', 'utf-8')
    #添加附件
    part = MIMEApplication(open(f'./monitor/数据统计周报-{date}.xlsx', 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename=f"数据统计周报-{date}.xlsx")
    msg.attach(part)
    # 消息体
    messageObj = MIMEText(message)
    msg.attach(messageObj)

    # html = """
    # <html>
    #   <head></head>
    #   <body>
    #     <p>Hi!<br>
    #        How are you?<br>
    #        Here is the <a href="http://www.baidu.com">link</a> you wanted.<br>
    #     </p>
    #   </body>
    # </html>
    # """
    # text_html = MIMEText(html, 'html', 'utf-8')
    # text_html["Content-Disposition"] = 'attachment; filename="texthtml.html"'
    # msg.attach(text_html)

    smtpObj = smtplib.SMTP("{}:{}".format('smtp-mail.outlook.com', 587))
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(sender, authCode)
    smtpObj.sendmail(sender, to_addrs=[email], msg=msg.as_string())
    smtpObj.quit()

def get_day(date, step=0):
    """获取指定日期date(形如"xxxx-xx-xx")之前或之后的多少天的日期, 返回值为字符串格式的日期"""
    l = date.split("-")
    y = int(l[0])
    m = int(l[1])
    d = int(l[2])
    old_date = datetime.datetime(y, m, d)
    new_date = (old_date + datetime.timedelta(days=step)).strftime('%Y-%m-%d')
    return new_date


def write_excel_xlsx(path, sheets):
    workbook = openpyxl.Workbook()
    count=0
    for i in sheets:
        value=i['values']
        index = len(value)
        sheet=workbook.create_sheet(i['sheet_name'],count)
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        count+=1
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")
    logger.info("xlsx格式表格写入数据成功！")


if __name__ == '__main__':
    """使用outlook发送。统计一周内清洗数据总量，并且发送邮件。
https://www.tablesgenerator.com/text_tables
+---+---+---+---+---+
| 1 | 2 | 4 | 3 | 4 |
+---+---+---+---+---+
|   |   |   |   |   |
+---+---+---+---+---+
|   |   |   |   |   |
+---+---+---+---+---+
|   |   |   |   |   |
+---+---+---+---+---+
    """
    time_array = time.localtime()
    now = time.strftime("%Y-%m-%d", time_array)
    create_time = get_day(now, 1)
    date = get_day(now, -7)
    logger.info(create_time)
    logger.info(date)
    connection = phoenix.connect_to_phoenix()
    phoenix_curs = connection.cursor()

    data_dict = {'CHA_BRANCH_BUSINESS': 0,
                 'CHA_BRANCH_BUS_STATION': 0,
                 'CHA_BRANCH_CREDITCARDARD': 0,
                 'CHA_BRANCH_FACILITY': 0,
                 'CHA_BRANCH_FINANCIAL_PRODUCT': 0,
                 'CHA_BRANCH_FUND_AGENCY': 0,
                 'CHA_BRANCH_FUND_BASIC': 0,
                 'CHA_BRANCH_HOSPITAL': 0,
                 'CHA_BRANCH_HOUSE': 0,
                 'CHA_BRANCH_HOUSE_DATA': 0,
                 'CHA_BRANCH_INSURANCE': 0,
                 'CHA_BRANCH_MAIN_ROUTE': 0,
                 'CHA_BRANCH_NEWS': 0,
                 'CHA_BRANCH_ORGANIZE': 0,
                 'CHA_BRANCH_SCHOOL': 0,
                 'CHA_BRANCH_SUBWAY': 0,
                 'CHA_BRANCH_WECHAT': 0,
                 'CHA_BRANCH_WEIBO_INFO': 0}

    #统计信息
    values = [["Hbase表名", "入库日期", "清洗总数"], ]
    phoenix_curs.execute(f"select * from (select sum(to_number(SHUFFLE_COUNT_)) "
                         f"as ct,HBASE_TABLE_ from CHA_DATA_MONITOR "
                         f"where DATE_>'{date}' and DATE_<'{create_time}' group by HBASE_TABLE_) i")
    all_data = phoenix_curs.fetchall()
    total=0
    for i in all_data:
        values.append([i[1], date+'到'+create_time, i[0]])
        total+=i[0]
        data_dict[i[1]] = int(i[0])
    values.append(['', '入库总量:', total])
    # 统计明细
    values1 = [["Hbase表名", "入库日期", "清洗总数"], ]
    phoenix_curs.execute(f"select * from (select sum(to_number(SHUFFLE_COUNT_)) "
                         f"as ct,HBASE_TABLE_,DATE_ from CHA_DATA_MONITOR "
                         f"where DATE_>'{date}' and DATE_<'{create_time}' group by HBASE_TABLE_,DATE_) i")
    all_data1 = phoenix_curs.fetchall()
    for i in all_data1:
        values1.append([i[1], i[2], i[0]])
    # 实体统计
    values2 = [["Hbase表名", "实体名称", "实体编码", "入库日期", "清洗总数"], ]
    phoenix_curs.execute(f"select HBASE_TABLE_,ENTITY_NAME_,ENTITY_CODE_,DATE_,SHUFFLE_COUNT_ from CHA_DATA_MONITOR where DATE_>'{date}' and DATE_<'{create_time}'")
    all_data2 = phoenix_curs.fetchall()
    for i in all_data2:
        values2.append([i[0], i[1], i[2],i[3],i[4]])

    connection.close()
    book_name_xlsx = f'./monitor/数据统计周报-{create_time}.xlsx'
    sheets=[]
    sheet0={'sheet_name':'清洗统计汇总','values':values}
    sheets.append(sheet0)

    sheet1={'sheet_name':'数据统计明细','values':values1}
    sheets.append(sheet1)

    sheet2={'sheet_name':'实体统计明细','values':values2}
    sheets.append(sheet2)

    write_excel_xlsx(book_name_xlsx, sheets)
    logger.info(data_dict)
    msg=f"""
各位尊敬的领导：
    本周数据清洗统计如下：{date} 到{create_time}
      1：财资：{data_dict['CHA_BRANCH_NEWS']}条
      2：微信：{data_dict['CHA_BRANCH_WECHAT']}条
      3：微博：{data_dict['CHA_BRANCH_WEIBO_INFO']}条
      4：网点：{data_dict['CHA_BRANCH_ORGANIZE']}条
      5：商圈：{data_dict['CHA_BRANCH_BUSINESS']}条
      6：学校：{data_dict['CHA_BRANCH_SCHOOL']}条
      7：医院：{data_dict['CHA_BRANCH_HOSPITAL']}条
      8：理财产品：{data_dict['CHA_BRANCH_FINANCIAL_PRODUCT']}条
      9：信用卡：{data_dict['CHA_BRANCH_CREDITCARDARD']}条
      10：基金：{data_dict['CHA_BRANCH_FUND_AGENCY']}条
      11：保险：{data_dict['CHA_BRANCH_INSURANCE']}条
      12：小区房价(链家)：{data_dict['CHA_BRANCH_HOUSE_DATA']}条
      13：周边设施：{data_dict['CHA_BRANCH_FACILITY']}条
      14：地铁：{data_dict['CHA_BRANCH_SUBWAY']}条
      15：公交：{data_dict['CHA_BRANCH_BUS_STATION']}条
      总计：{total}
      详情请见附件！
  谢谢查收!
        """
    sendEmailMessage('wenzhong.ding@pactera.com', msg, create_time)
    sendEmailMessage('1059792930@qq.com', msg, create_time)
